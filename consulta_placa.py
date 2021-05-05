from openpyxl import Workbook, load_workbook
from selenium import webdriver
import string
import time
from unidecode import unidecode

estados = {"11": "RO",
           "12": "AC",
           "13": "AM",
           "14": "RR",
           "15": "PA",
           "16": "AP",
           "17": "TO",
           "21": "MA",
           "22": "PI",
           "23": "CE",
           "24": "RN",
           "25": "PB",
           "26": "PE",
           "27": "AL",
           "28": "SE",
           "29": "BA",
           "31": "MG",
           "32": "ES",
           "33": "RJ",
           "35": "SP",
           "41": "PR",
           "42": "SC",
           "43": "RS",
           "50": "MS",
           "51": "MT",
           "52": "GO",
           "53": "DF", }


def procurar_estado_por_codigo(codigo_estado):
    return estados[codigo_estado]


def procurar_codigo_estado_por_uf(uf):
    for chave, valor in estados.items():
        if uf.upper() == valor.upper():
            return chave


def remover_caracteres(texto):
    novo_texto = ''.join(e for e in texto if e.isalnum()).lower()
    return unidecode(novo_texto)


class Placa:
    def __init__(self):
        self.placa = ""
        self.uf = ""
        self.cidade = ""
        self.codigo = ""


class ConsultaPlaca:

    @staticmethod
    def consulta_placas(placas):
        browser = webdriver.Chrome()
        novas_placas = []
        for placa in placas:
            browser.get("http://infocarrosp.com.br/")
            browser.find_element_by_xpath('//*[@id="plate"]').send_keys(placa)
            browser.find_element_by_xpath('/html/body/div/div[1]/div/div[1]/form/div[2]/div/button').click()
            time.sleep(5)
            resp = browser.find_element_by_xpath('/html/body/div/div[1]/div[3]/table/tbody/tr[2]/td').text
            res = string.capwords(resp)
            res = res.split("/")
            cidade = res[0].strip()
            uf = res[1].strip()

            # criando uma nova placa com os dados obtidos
            nova_placa = Placa()
            nova_placa.placa = placa
            nova_placa.uf = uf
            nova_placa.cidade = cidade

            novas_placas.append(nova_placa)

        return novas_placas


def percorrer_planilha(placa):
    uf_procurada = placa.uf
    codigo_uf_procurada = procurar_codigo_estado_por_uf(uf_procurada)
    cidade_procurada = remover_caracteres(placa.cidade)

    arquivo_excel = load_workbook("municipios.xlsx")
    planilha1 = arquivo_excel.active

    max_linha = planilha1.max_row
    for i in range(2, max_linha + 1):
        codigo_uf_arquivo = planilha1.cell(row=i, column=1).value
        cidade_arquivo = planilha1.cell(row=i, column=3).value
        cidade_arquivo = remover_caracteres(cidade_arquivo)
        if codigo_uf_arquivo == codigo_uf_procurada \
                and cidade_procurada == cidade_arquivo:
            codigo_cidade = planilha1.cell(row=i, column=2).value
            placa.codigo = "{} {}".format(uf_procurada.upper(), codigo_cidade)


placass = ['FFX1128', 'HMV1135', 'PUB1179', 'MMX1530']
consulta = ConsultaPlaca.consulta_placas(placass)

for placa in consulta:
    percorrer_planilha(placa)

