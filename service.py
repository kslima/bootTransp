import xml.etree.ElementTree as Et

from openpyxl import load_workbook
from unidecode import unidecode

from model import Motorista, Produto, Veiculo, Municipio

__estados = {"11": "RO",
             "12": "AC",
             "13": "AM",
             "14": "RR",
             "15": "PA",
             "16": "AP",
             "17": "TO",
             "21": "MA",
             "22": "PI",
             "23": "CE",
             "24": "RN",
             "25": "PB",
             "26": "PE",
             "27": "AL",
             "28": "SE",
             "29": "BA",
             "31": "MG",
             "32": "ES",
             "33": "RJ",
             "35": "SP",
             "41": "PR",
             "42": "SC",
             "43": "RS",
             "50": "MS",
             "51": "MT",
             "52": "GO",
             "53": "DF", }

FILE_PATH = "properties.xml"


def load_xml_file():
    source = open(FILE_PATH)
    tree = Et.parse(source)
    root = tree.getroot()
    return source, tree, root


def listar_produtos():
    xml = load_xml_file()
    root = xml[2]
    produtos = []
    for tag_produto in root.findall("produto"):
        for item in tag_produto.findall("item"):
            produto = Produto()
            produto.codigo = item.get("codigo")
            produto.nome = item.get("nome")
            produto.deposito = item.get("deposito")
            produto.lote = item.get("lote")
            produto.inspecao_veiculo = item.get("inspecao_veiculo")
            produto.inspecao_produto = item.get("inspecao_produto")
            produto.remover_a = item.get("remover_a")
            produtos.append(produto)
    xml[0].close()
    return produtos


# pesquisa um produto pela descricao
def procurar_produto_pelo_nome(nome_produto):
    produtos = listar_produtos()
    for p in produtos:
        if p.nome == nome_produto:
            return p


# pesquisa um produto pela descricao
def procurar_produto_pelo_codigo(codigo_produto):
    produtos = listar_produtos()
    for p in produtos:
        if p.codigo == codigo_produto:
            return p


def cadastrar_produto_se_nao_exister(novo_produto):
    xml = load_xml_file()
    root = xml[2]
    produto_procurado = procurar_produto_pelo_codigo(novo_produto.codigo)
    if produto_procurado is None:
        try:
            atributos_novo_produto = {"codigo": novo_produto.codigo,
                                      "nome": novo_produto.nome,
                                      "deposito": novo_produto.deposito,
                                      "lote": novo_produto.lote,
                                      "inspecao_veiculo": novo_produto.inspecao_veiculo,
                                      "inspecao_produto": novo_produto.inspecao_produto,
                                      "remover_a": novo_produto.remover_a}

            produto_tag = root.find('produto')
            item = Et.SubElement(produto_tag, 'item', atributos_novo_produto)
            item.attrib = atributos_novo_produto
            xml[1].write(FILE_PATH)
            return 1, "Produto cadastrado com sucesso!"
        except Exception as e:
            print(e)
            return 0, "Erro ao cadastrar novo produto!\n{}".format(str(e))
        finally:
            xml[0].close()
    else:
        return -1, "Já existe um produto cadastrado com esses dados!"


def atualizar_produto(produto_para_atualizar):
    xml = load_xml_file()
    try:
        root = xml[2]
        for produto_tag in root.findall("produto"):
            for item in produto_tag.findall("item"):
                if item.get("codigo") == produto_para_atualizar.codigo:
                    item.attrib['nome'] = produto_para_atualizar.nome
                    item.attrib['deposito'] = produto_para_atualizar.deposito
                    item.attrib['lote'] = produto_para_atualizar.lote
                    item.attrib['inspecao_veiculo'] = produto_para_atualizar.inspecao_veiculo
                    item.attrib['inspecao_produto'] = produto_para_atualizar.inspecao_produto
                    item.attrib['remover_a'] = produto_para_atualizar.remover_a
        xml[1].write(FILE_PATH)
        return True, "Produto '{}' atualizado com sucesso!".format(produto_para_atualizar.codigo)
    except Exception as e:
        return False, "Erro ao atualizar novo produto!\n{}".format(str(e))
    finally:
        xml[0].close()


# lista todos os motoristas
def listar_motoristas():
    xml = load_xml_file()
    root = xml[2]
    motoristas = []
    for tag_motorista in root.findall("motorista"):
        for item in tag_motorista.findall("item"):
            motorista = Motorista()
            motorista.nome = item.get("nome")
            motorista.cpf = item.get("cpf")
            motorista.cnh = item.get("cnh")
            motorista.rg = item.get("rg")
            motoristas.append(motorista)
    xml[0].close()
    return motoristas


# pesquisa um motorista pelo cpf, cnh ou rg
def procurar_motorista_por_documento(*args):
    motoristas = listar_motoristas()
    _motorista = None
    for motorista in motoristas:
        for documento in args:
            if __comparar_ignorando_vazio(motorista.cpf, documento) or \
                    __comparar_ignorando_vazio(motorista.cnh, documento) or \
                    __comparar_ignorando_vazio(motorista.rg, documento):
                _motorista = motorista

    # caso nao encontre, ele verifica se contem
    if _motorista is None:
        for motorista in motoristas:
            for documento in args:
                if __verificar_se_contem_ignorando_vazio(documento, motorista.cpf) or \
                        __verificar_se_contem_ignorando_vazio(documento, motorista.cnh) or \
                        __verificar_se_contem_ignorando_vazio(documento, motorista.rg):
                    _motorista = motorista
    return _motorista


# cria um novo motorista
def cadastrar_motorista_se_nao_existir(novo_motorista):
    xml = load_xml_file()
    root = xml[2]
    motorista_procurado = procurar_motorista_por_documento(novo_motorista.cpf, novo_motorista.cnh, novo_motorista.rg)
    if motorista_procurado is None:
        try:
            atributos_motorista = {"nome": novo_motorista.nome,
                                   "cpf": novo_motorista.cpf,
                                   "cnh": novo_motorista.cnh,
                                   "rg": novo_motorista.rg}

            motorista_tag = root.find('motorista')
            item = Et.SubElement(motorista_tag, 'item', atributos_motorista)
            item.attrib = atributos_motorista
            xml[1].write(FILE_PATH)
            return 1, "Motorista cadastrado com sucesso!"
        except Exception as e:
            print(e)
            return 0, "Erro ao cadastrar novo motorista!\n{}".format(str(e))
        finally:
            xml[0].close()
    else:
        return -1, "já existe um motorista cadastrado com esses dados!"


def atualizar_motorista(motorista_para_atualizar):
    xml = load_xml_file()
    try:
        root = xml[2]
        for tag_motorista in root.findall("motorista"):
            for item in tag_motorista.findall("item"):
                if item.get("cpf") == motorista_para_atualizar.cpf or item.get("cnh") == motorista_para_atualizar.cnh \
                        or item.get("rg") == motorista_para_atualizar.rg:
                    item.attrib['nome'] = motorista_para_atualizar.nome
                    item.attrib['cpf'] = motorista_para_atualizar.cpf
                    item.attrib['cnh'] = motorista_para_atualizar.cnh
                    item.attrib['rg'] = motorista_para_atualizar.rg

        xml[1].write(FILE_PATH)
        return True, "Motorista '{}' atualizado com sucesso!".format(motorista_para_atualizar.nome)
    except Exception as e:
        return False, "Erro ao atualizar motorista {}!\n{}".format(motorista_para_atualizar.nome, str(e))
    finally:
        xml[0].close()


def listar_veiculos():
    xml = load_xml_file()
    root = xml[2]
    veiculos = []
    for tag_veiculo in root.findall("veiculo"):
        for item in tag_veiculo.findall("item"):
            veiculo = Veiculo()
            veiculo.tipo_veiculo = item.get("tipo_veiculo")
            veiculo.tolerancia_balanca = item.get("tolerancia_balanca")
            veiculo.quantidade_lacres = item.get("quantidade_lacres")
            veiculo.placa_1 = item.get("placa_1")
            veiculo.placa_2 = item.get("placa_2")
            veiculo.placa_3 = item.get("placa_3")
            veiculo.placa_4 = item.get("placa_4")
            veiculo.codigo_municipio_placa_1 = item.get("codigo_municipio_placa_1")
            veiculo.codigo_municipio_placa_2 = item.get("codigo_municipio_placa_2")
            veiculo.codigo_municipio_placa_3 = item.get("codigo_municipio_placa_3")
            veiculo.codigo_municipio_placa_4 = item.get("codigo_municipio_placa_4")
            veiculos.append(veiculo)
    xml[0].close()
    return veiculos


def procurar_veiculos(placa):
    lista_veiculos = listar_veiculos()
    veiculos_encontrados = []
    for veiculo in lista_veiculos:
        if veiculo.placa_1 == placa:
            veiculos_encontrados.append(veiculo)
    return veiculos_encontrados


def cadastrar_veiculo_se_nao_exister(novo_veiculo):
    xml = load_xml_file()
    root = xml[2]
    veiculo_procurado = procurar_veiculos(novo_veiculo.placa_1)
    if len(veiculo_procurado) == 0:
        try:
            atributos_novo_produto = {"tipo_veiculo": novo_veiculo.tipo_veiculo.strip(),
                                      "tolerancia_balanca": novo_veiculo.tolerancia_balanca.strip(),
                                      "quantidade_lacres": novo_veiculo.quantidade_lacres.strip(),
                                      "placa_1": novo_veiculo.placa_1.strip(),
                                      "placa_2": novo_veiculo.placa_2.strip(),
                                      "placa_3": novo_veiculo.placa_3.strip(),
                                      "placa_4": novo_veiculo.placa_4.strip(),
                                      "codigo_municipio_placa_1": novo_veiculo.codigo_municipio_placa_1.strip(),
                                      "codigo_municipio_placa_2": novo_veiculo.codigo_municipio_placa_2.strip(),
                                      "codigo_municipio_placa_3": novo_veiculo.codigo_municipio_placa_3.strip(),
                                      "codigo_municipio_placa_4": novo_veiculo.codigo_municipio_placa_4.strip()}

            tag_veiculo = root.find('veiculo')
            item = Et.SubElement(tag_veiculo, 'item', atributos_novo_produto)
            item.attrib = atributos_novo_produto
            xml[1].write(FILE_PATH)
            return 1, "Veículo cadastrado com sucesso!"
        except Exception as e:
            print(e)
            return 0, "Erro ao cadastrar novo Veículo!\n{}".format(str(e))
        finally:
            xml[0].close()
    else:
        return -1, "Já existe um veículo cadastrado com esses dados!"


def atualizar_veiculo(veiculo_para_atualizar):
    xml = load_xml_file()
    try:
        root = xml[2]
        for tag_veiculo in root.findall("veiculo"):
            for item in tag_veiculo.findall("item"):
                if item.get("placa_1") == veiculo_para_atualizar.placa_1:
                    item.attrib['tipo_veiculo'] = veiculo_para_atualizar.tipo_veiculo
                    item.attrib['tolerancia_balanca'] = veiculo_para_atualizar.tolerancia_balanca
                    item.attrib['quantidade_lacres'] = veiculo_para_atualizar.quantidade_lacres
                    item.attrib['placa_1'] = veiculo_para_atualizar.placa_1
                    item.attrib['placa_2'] = veiculo_para_atualizar.placa_2
                    item.attrib['placa_3'] = veiculo_para_atualizar.placa_3
                    item.attrib['placa_4'] = veiculo_para_atualizar.placa_4
                    item.attrib['codigo_municipio_placa_1'] = veiculo_para_atualizar.codigo_municipio_placa_1
                    item.attrib['codigo_municipio_placa_2'] = veiculo_para_atualizar.codigo_municipio_placa_2
                    item.attrib['codigo_municipio_placa_3'] = veiculo_para_atualizar.codigo_municipio_placa_3
                    item.attrib['codigo_municipio_placa_4'] = veiculo_para_atualizar.codigo_municipio_placa_4

        xml[1].write(FILE_PATH)
        return True, "Veículo '{}' atualizado com sucesso!".format(veiculo_para_atualizar.placa_1)
    except Exception as e:
        return False, "Erro ao atualizar veículo {}!\n{}".format(veiculo_para_atualizar.placa_1, str(e))
    finally:
        xml[0].close()


def __comparar_ignorando_vazio(v1, v2):
    if v1 == "" or v2 == "":
        return False
    return v1 == v2


def __verificar_se_contem_ignorando_vazio(v1, v2):
    if v1 == "" or v2 == "":
        return False
    return v1 in v2


def listar_municipios_brasileiros():
    arquivo_excel = load_workbook("municipios.xlsx")
    planilha1 = arquivo_excel.active
    max_linha = planilha1.max_row

    municipios = []
    for i in range(2, max_linha + 1):
        codigo_uf = planilha1.cell(row=i, column=1).value
        codigo_municipio = planilha1.cell(row=i, column=2).value
        municipio = planilha1.cell(row=i, column=3).value
        municipios.append(Municipio(codigo_uf, municipio, codigo_municipio))
    return municipios


def __procurar_codigo_estado_por_uf(uf):
    for chave, valor in __estados.items():
        if uf.upper() == valor.upper():
            return chave


def __remover_caracteres(texto):
    novo_texto = ''.join(e for e in texto if e.isalnum()).lower()
    return unidecode(novo_texto)


def listar_tolerancias_balanca():
    return ["Z2 - Vei. 2 eix(16.000)",
            "Z3 - Vei. 3 eix(23.000)",
            "Z4 - Vei. 4 eix < 16m(31500)",
            "Z5 - Vei. 5 eix < 16m(41500)",
            "Z6 - Vei. 5 eix < 16m distan(45000)",
            "Z7 - Vei. 5 eix >= 16m(41500)",
            "Z8 - Vei. 6 eix < 16m(45000)",
            "Z9 - Vei. 6 eix >= 16m tandem(48500)",
            "ZA - Vei. 6 eix >= 16m tandem + Isol(50.000)",
            "ZB - Vei. 6 eix >= 16m distan(53.000)",
            "ZC - Vei. 7 eix(57.000)",
            "ZD - Vei. 9 eix(74.000)",
            "ZE - Vei. 7 eix Tq c/ Lic(59.850)",
            "ZF - Vei. 9 eix Tq c/ Lic(77.700)",
            "ZG - Vei. 5 eix > 16m distan(46.000)",
            "ZH - Vei. 6 eix >= 16m distan c/ Lic(55.650)",
            "ZI - Vei. 5 eix > 16m C/ Licença(43.580)",
            "ZJ - Vei. 3 eix Cav.Mec + SemiReboq.(26.000)",
            "ZK - Vei. 5 eix Cav.Mec + SemiReboq.(40.000)",
            "ZL - Vei. 3 eix Cav.Mec+Semi c/lic.(27.300)"]


def listar_tipos_veiculos():
    return ["01 - Outros",
            "02 - Postal",
            "03 - Ferroviário",
            "04 - Marítimo",
            "05 - Graneleiro",
            "06 - Caçamba",
            "07 - Tanque",
            "08 - Coleta",
            "09 - Bi Caçamba",
            "10 - Bi Graneleiro",
            "11 - Hopper",
            "12 - Bi Tanque",
            "13 - RodoTrem",
            "14 - Vanderleia",
            "15 - PICK-UP",
            "16 - VEICULO 3/4",
            "17 - TRUCK",
            "18 - CARRETA",
            "19 - Bi-Trem"]
